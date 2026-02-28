# build_suite.py
# Deterministic generator for: 4 artwork masters + zonemaps + workbook + access schema + manifest + zip
# Spec: INFUSIST LAB — EXPERIMENTAL ART SUITE (D) — SPEC v1.0
#
# Safety rules honored:
# - Originals immutable (only writes to out_dir)
# - Deterministic outputs (no external randomness; fixed seed + fixed transforms)
# - No silent drift (manifest with SHA256 + settings_json + generator_version)

from __future__ import annotations

import argparse
import hashlib
import io
import json
import math
import os
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Dict, List, Tuple

import numpy as np
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


GENERATOR_VERSION = "exp-art-suite-d-builder-1.0.0"
SUITE_VERSION = "1.0"
ARTWORK_VERSION = "1.0.0"
ZONEMAP_VERSION = "1.0.0"

# Canonical master size (can be overridden by CLI)
DEFAULT_W = 3000
DEFAULT_H = 2000

# Edge ring: 3% of shortest dimension at Y=3% luminance
EDGE_RING_FRAC = 0.03
EDGE_RING_Y_PCT = 3  # percent luminance target


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def y_pct_to_srgb8(y_pct: float) -> int:
    """
    Convert target relative luminance (approx) to an sRGB 8-bit grayscale value.
    IMPORTANT: The spec references "linear-mapped workflow"; in production, you'd use
    a proper color-managed pipeline. For deterministic test art, we define a fixed mapping:

    - We interpret y_pct as linear light percentage (0..100).
    - Convert linear to sRGB with standard gamma approx:
      sRGB = 12.92*L for L<=0.0031308 else 1.055*L^(1/2.4)-0.055
    - Return 8-bit integer.

    This is deterministic and documented. ZoneMap stores both y_pct and resulting srgb8.
    """
    L = max(0.0, min(1.0, y_pct / 100.0))
    if L <= 0.0031308:
        s = 12.92 * L
    else:
        s = 1.055 * (L ** (1 / 2.4)) - 0.055
    v = int(round(max(0.0, min(1.0, s)) * 255.0))
    return max(0, min(255, v))


def clamp01(x: np.ndarray) -> np.ndarray:
    return np.clip(x, 0.0, 1.0)


def linear_to_srgb8(arr_lin: np.ndarray) -> np.ndarray:
    # arr_lin: 0..1 linear grayscale
    a = arr_lin.copy()
    out = np.empty_like(a)
    m = a <= 0.0031308
    out[m] = 12.92 * a[m]
    out[~m] = 1.055 * np.power(a[~m], 1 / 2.4) - 0.055
    out = clamp01(out)
    return np.round(out * 255.0).astype(np.uint8)


def make_canvas(w: int, h: int, base_y_pct: float) -> np.ndarray:
    base = np.full((h, w), base_y_pct / 100.0, dtype=np.float32)
    return base


def apply_edge_ring_lin(canvas: np.ndarray, ring_px: int, y_pct: float) -> None:
    h, w = canvas.shape
    L = y_pct / 100.0
    # Top
    canvas[0:ring_px, :] = L
    # Bottom
    canvas[h - ring_px : h, :] = L
    # Left
    canvas[:, 0:ring_px] = L
    # Right
    canvas[:, w - ring_px : w] = L


def draw_rect_lin(canvas: np.ndarray, x0: int, y0: int, x1: int, y1: int, y_pct: float) -> None:
    canvas[y0:y1, x0:x1] = y_pct / 100.0


def draw_gradient_h_lin(canvas: np.ndarray, x0: int, y0: int, x1: int, y1: int, y0_pct: float, y1_pct: float) -> None:
    width = x1 - x0
    if width <= 0:
        return
    grad = np.linspace(y0_pct / 100.0, y1_pct / 100.0, width, dtype=np.float32)
    canvas[y0:y1, x0:x1] = grad[None, :]


def draw_gradient_v_lin(canvas: np.ndarray, x0: int, y0: int, x1: int, y1: int, y0_pct: float, y1_pct: float) -> None:
    height = y1 - y0
    if height <= 0:
        return
    grad = np.linspace(y0_pct / 100.0, y1_pct / 100.0, height, dtype=np.float32)
    canvas[y0:y1, x0:x1] = grad[:, None]


def add_microtexture(canvas: np.ndarray, x0: int, y0: int, x1: int, y1: int, base_y_pct: float, modulation_y_pct: float) -> None:
    """
    1px modulation checker for microtexture.
    base_y_pct +/- modulation_y_pct/2 in linear light.
    """
    h = y1 - y0
    w = x1 - x0
    if h <= 0 or w <= 0:
        return
    base = base_y_pct / 100.0
    mod = modulation_y_pct / 100.0
    yy, xx = np.mgrid[0:h, 0:w]
    pattern = ((xx + yy) & 1).astype(np.float32)  # 0/1
    vals = base + (pattern - 0.5) * mod
    canvas[y0:y1, x0:x1] = clamp01(vals)


def save_png_grayscale(path: str, canvas_lin: np.ndarray) -> None:
    img = Image.fromarray(linear_to_srgb8(canvas_lin), mode="L").convert("RGB")
    img.save(path, format="PNG", optimize=True)


def save_png_rgb(path: str, arr_rgb_srgb8: np.ndarray) -> None:
    img = Image.fromarray(arr_rgb_srgb8.astype(np.uint8), mode="RGB")
    img.save(path, format="PNG", optimize=True)


def build_art_01(out_dir: str, w: int, h: int) -> Tuple[str, Dict]:
    """
    ART-01: TONAL STRESS FIELD (Shadows/Highlights/Banding)
    - Minimal monochrome blur aesthetic (we emulate with smooth gradients).
    - Embedded tests: shadow ladder, microtexture, midtone ramp, highlight steps, edge ring.
    """
    ring_px = int(round(min(w, h) * EDGE_RING_FRAC))
    # base field: dark-to-mid smooth gradient to simulate depth blur
    canvas = make_canvas(w, h, base_y_pct=12.0)
    # big soft vertical gradient: 8% at top to 22% at bottom
    draw_gradient_v_lin(canvas, 0, 0, w, h, 8.0, 22.0)

    zones = {}

    # Z1 Shadow Ladder: 0,2,4,6,8,10% luminance steps
    z1 = {
        "id": "Z1",
        "name": "Shadow Ladder",
        "steps_y_pct": [0, 2, 4, 6, 8, 10],
    }
    x0 = int(w * 0.10)
    y0 = int(h * 0.15)
    x1 = int(w * 0.90)
    band_h = int(h * 0.01)  # micro-bands
    gap = int(h * 0.004)
    y = y0
    for s in z1["steps_y_pct"]:
        draw_rect_lin(canvas, x0, y, x1, y + band_h, s)
        y += band_h + gap
    z1["rect"] = [x0, y0, x1, y - gap]
    zones["Z1"] = z1

    # Z2 Shadow Microtexture: 1px modulation at 3%,5%,7% luminance
    z2 = {"id": "Z2", "name": "Shadow Microtexture", "patches": []}
    pxw = int(w * 0.18)
    pxh = int(h * 0.08)
    base_levels = [3, 5, 7]
    mod = 0.8  # small modulation in Y% to be subtle but measurable
    start_x = int(w * 0.10)
    start_y = int(h * 0.33)
    pad = int(w * 0.02)
    for i, base in enumerate(base_levels):
        xx0 = start_x + i * (pxw + pad)
        yy0 = start_y
        xx1 = xx0 + pxw
        yy1 = yy0 + pxh
        add_microtexture(canvas, xx0, yy0, xx1, yy1, base_y_pct=base, modulation_y_pct=mod)
        z2["patches"].append({"rect": [xx0, yy0, xx1, yy1], "base_y_pct": base, "modulation_y_pct": mod})
    # envelope rect
    z2["rect"] = [start_x, start_y, start_x + 3 * pxw + 2 * pad, start_y + pxh]
    zones["Z2"] = z2

    # Z3 Midtone Ramp: 30–70% gradient strip
    x0 = int(w * 0.10)
    y0 = int(h * 0.50)
    x1 = int(w * 0.90)
    y1 = int(h * 0.58)
    draw_gradient_h_lin(canvas, x0, y0, x1, y1, 30.0, 70.0)
    zones["Z3"] = {"id": "Z3", "name": "Midtone Ramp", "rect": [x0, y0, x1, y1], "y_pct_range": [30, 70]}

    # Z4 Highlight Steps: 92–100% with 1% separations
    x0 = int(w * 0.10)
    y0 = int(h * 0.65)
    x1 = int(w * 0.90)
    y1 = int(h * 0.80)
    step_vals = list(range(92, 101))
    step_w = max(1, (x1 - x0) // len(step_vals))
    xx = x0
    for s in step_vals:
        draw_rect_lin(canvas, xx, y0, min(x1, xx + step_w), y1, float(s))
        xx += step_w
    zones["Z4"] = {"id": "Z4", "name": "Highlight Steps", "rect": [x0, y0, x1, y1], "steps_y_pct": step_vals}

    # Z5 Edge ring
    apply_edge_ring_lin(canvas, ring_px, EDGE_RING_Y_PCT)
    zones["Z5"] = {"id": "Z5", "name": "Peripheral Mount Band", "ring_px": ring_px, "y_pct": EDGE_RING_Y_PCT}

    out_path = os.path.join(out_dir, f"ART-01_TonalStressField_v{ARTWORK_VERSION}.png")
    save_png_grayscale(out_path, canvas)

    zonemap = {
        "Artwork_ID": "ART-01",
        "Artwork_Version": ARTWORK_VERSION,
        "ZoneMap_Version": ZONEMAP_VERSION,
        "MasterResolution": {"w": w, "h": h},
        "EdgeRing": {"width_px": ring_px, "y_pct": EDGE_RING_Y_PCT, "srgb8": y_pct_to_srgb8(EDGE_RING_Y_PCT)},
        "Zones": zones,
        "Notes": "Deterministic test master; grayscale; embedded zones per Spec v1.0",
    }
    return out_path, zonemap


def build_art_02(out_dir: str, w: int, h: int) -> Tuple[str, Dict]:
    """
    ART-02: MIDTONE PRESERVATION FIELD (Midtone density / Pop)
    Embedded: midtone step blocks, micro-contrast patch, diagonal ramp, edge ring.
    """
    ring_px = int(round(min(w, h) * EDGE_RING_FRAC))
    canvas = make_canvas(w, h, base_y_pct=50.0)

    # Fog aesthetic: gentle radial falloff (deterministic)
    yy, xx = np.mgrid[0:h, 0:w]
    cx, cy = w * 0.5, h * 0.5
    r = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2) / np.sqrt(cx**2 + cy**2)
    fog = 0.50 + (0.06 * (1.0 - r)).astype(np.float32)  # 50% +/- 6%
    canvas = clamp01(fog)

    zones = {}

    # Z1 Midtone Step Blocks: 40..65% (6 blocks)
    steps = [40, 45, 50, 55, 60, 65]
    x0 = int(w * 0.08)
    y0 = int(h * 0.15)
    x1 = int(w * 0.92)
    y1 = int(h * 0.30)
    block_w = (x1 - x0) // len(steps)
    xx0 = x0
    for s in steps:
        draw_rect_lin(canvas, xx0, y0, xx0 + block_w, y1, float(s))
        xx0 += block_w
    zones["Z1"] = {"id": "Z1", "name": "Midtone Step Blocks", "rect": [x0, y0, x1, y1], "steps_y_pct": steps}

    # Z2 Micro-Contrast Patch: ±2 luminance modulation embedded in 50–55% area
    # We'll create a subtle sine modulation around 52.5% with amplitude 2%.
    x0 = int(w * 0.10)
    y0 = int(h * 0.36)
    x1 = int(w * 0.40)
    y1 = int(h * 0.48)
    patch_h = y1 - y0
    patch_w = x1 - x0
    yy2, xx2 = np.mgrid[0:patch_h, 0:patch_w]
    base = 0.525
    amp = 0.02
    mod = base + amp * np.sin(2 * math.pi * xx2 / 12.0)  # 12px period deterministic
    canvas[y0:y1, x0:x1] = clamp01(mod.astype(np.float32))
    zones["Z2"] = {"id": "Z2", "name": "Micro-Contrast Patch", "rect": [x0, y0, x1, y1], "base_y_pct": 52.5, "delta_y_pct": 2.0}

    # Z3 Diagonal Ramp: 35–70% smooth gradient
    x0 = int(w * 0.55)
    y0 = int(h * 0.36)
    x1 = int(w * 0.92)
    y1 = int(h * 0.62)
    hh = y1 - y0
    ww = x1 - x0
    yy3, xx3 = np.mgrid[0:hh, 0:ww]
    t = (xx3 / max(1, ww - 1) + yy3 / max(1, hh - 1)) * 0.5
    ramp = (0.35 + (0.70 - 0.35) * t).astype(np.float32)
    canvas[y0:y1, x0:x1] = clamp01(ramp)
    zones["Z3"] = {"id": "Z3", "name": "Diagonal Ramp", "rect": [x0, y0, x1, y1], "y_pct_range": [35, 70]}

    # Z4 Edge ring
    apply_edge_ring_lin(canvas, ring_px, EDGE_RING_Y_PCT)
    zones["Z4"] = {"id": "Z4", "name": "Peripheral Mount Band", "ring_px": ring_px, "y_pct": EDGE_RING_Y_PCT}

    out_path = os.path.join(out_dir, f"ART-02_MidtonePreservationField_v{ARTWORK_VERSION}.png")
    save_png_grayscale(out_path, canvas)

    zonemap = {
        "Artwork_ID": "ART-02",
        "Artwork_Version": ARTWORK_VERSION,
        "ZoneMap_Version": ZONEMAP_VERSION,
        "MasterResolution": {"w": w, "h": h},
        "EdgeRing": {"width_px": ring_px, "y_pct": EDGE_RING_Y_PCT, "srgb8": y_pct_to_srgb8(EDGE_RING_Y_PCT)},
        "Zones": zones,
        "Notes": "Deterministic test master; grayscale; embedded zones per Spec v1.0",
    }
    return out_path, zonemap


def build_art_03(out_dir: str, w: int, h: int) -> Tuple[str, Dict]:
    """
    ART-03: COLOR NEUTRALITY MATRIX (Cast / Skin proxy / Saturation)
    RGB image with neutral reference and controlled color fields.
    """
    ring_px = int(round(min(w, h) * EDGE_RING_FRAC))

    # Base: muted premium minimal color field background
    img = np.zeros((h, w, 3), dtype=np.uint8)
    img[:, :] = np.array([110, 112, 114], dtype=np.uint8)  # slightly cool gray background

    zones = {}

    # Z1 Neutral Reference: RGB 128/128/128 block
    x0 = int(w * 0.08)
    y0 = int(h * 0.12)
    x1 = int(w * 0.30)
    y1 = int(h * 0.28)
    img[y0:y1, x0:x1] = np.array([128, 128, 128], dtype=np.uint8)
    zones["Z1"] = {"id": "Z1", "name": "Neutral Reference", "rect": [x0, y0, x1, y1], "rgb": [128, 128, 128]}

    # Z2 Channel Isolators: low-sat bands at 20% intensity (approx)
    # We'll set bands with one channel elevated slightly while keeping low saturation.
    band_y0 = int(h * 0.32)
    band_y1 = int(h * 0.42)
    band_h = band_y1 - band_y0
    band_w = int(w * 0.20)
    gap = int(w * 0.03)
    start_x = int(w * 0.08)
    # R-only (low sat): raise R by 20% of range around mid gray
    bands = [
        ("R", [150, 120, 120]),
        ("G", [120, 150, 120]),
        ("B", [120, 120, 150]),
    ]
    patches = []
    for i, (name, rgb) in enumerate(bands):
        xx0 = start_x + i * (band_w + gap)
        xx1 = xx0 + band_w
        img[band_y0:band_y1, xx0:xx1] = np.array(rgb, dtype=np.uint8)
        patches.append({"channel": name, "rect": [xx0, band_y0, xx1, band_y1], "rgb": rgb})
    zones["Z2"] = {"id": "Z2", "name": "Channel Isolators", "patches": patches, "intensity_note": "low-sat approx 20% bias"}

    # Z3 Skin Proxy Ramp: warm hue band (approx hue 25–35°) 45–65% luminance
    # Implement as horizontal gradient between two warm tones.
    x0 = int(w * 0.08)
    y0 = int(h * 0.48)
    x1 = int(w * 0.55)
    y1 = int(h * 0.62)
    ww = x1 - x0
    left = np.array([170, 130, 100], dtype=np.float32)  # darker warm
    right = np.array([210, 165, 135], dtype=np.float32)  # lighter warm
    grad = np.linspace(0.0, 1.0, ww, dtype=np.float32)[None, :, None]
    ramp = (left[None, None, :] * (1 - grad) + right[None, None, :] * grad).astype(np.uint8)
    img[y0:y1, x0:x1] = np.repeat(ramp, y1 - y0, axis=0)
    zones["Z3"] = {"id": "Z3", "name": "Skin Proxy Ramp", "rect": [x0, y0, x1, y1], "note": "warm band approx hue 25–35°, 45–65% luminance proxy"}

    # Z4 Saturation Ladder: 20/40/60/80% chroma blocks (approx)
    # Use a base hue and vary saturation by mixing with gray.
    x0 = int(w * 0.62)
    y0 = int(h * 0.12)
    x1 = int(w * 0.92)
    y1 = int(h * 0.28)
    steps = [20, 40, 60, 80]
    block_w = (x1 - x0) // len(steps)
    base_color = np.array([70, 140, 190], dtype=np.float32)  # muted teal-ish
    gray = np.array([128, 128, 128], dtype=np.float32)
    xx0 = x0
    blocks = []
    for s in steps:
        t = s / 100.0
        rgb = (gray * (1 - t) + base_color * t).round().astype(np.uint8).tolist()
        xx1 = xx0 + block_w
        img[y0:y1, xx0:xx1] = np.array(rgb, dtype=np.uint8)
        blocks.append({"chroma_pct": s, "rect": [xx0, y0, xx1, y1], "rgb": rgb})
        xx0 = xx1
    zones["Z4"] = {"id": "Z4", "name": "Saturation Ladder", "blocks": blocks, "note": "chroma via gray mix; deterministic"}

    # Edge ring: exact Y=3% is grayscale; we set RGB all equal to srgb8(Y=3%).
    ring_val = y_pct_to_srgb8(EDGE_RING_Y_PCT)
    img[0:ring_px, :, :] = ring_val
    img[h - ring_px : h, :, :] = ring_val
    img[:, 0:ring_px, :] = ring_val
    img[:, w - ring_px : w, :] = ring_val
    zones["Z5"] = {"id": "Z5", "name": "Peripheral Mount Band", "ring_px": ring_px, "y_pct": EDGE_RING_Y_PCT, "rgb": [ring_val, ring_val, ring_val]}

    out_path = os.path.join(out_dir, f"ART-03_ColorNeutralityMatrix_v{ARTWORK_VERSION}.png")
    save_png_rgb(out_path, img)

    zonemap = {
        "Artwork_ID": "ART-03",
        "Artwork_Version": ARTWORK_VERSION,
        "ZoneMap_Version": ZONEMAP_VERSION,
        "MasterResolution": {"w": w, "h": h},
        "EdgeRing": {"width_px": ring_px, "y_pct": EDGE_RING_Y_PCT, "srgb8": ring_val},
        "Zones": zones,
        "Notes": "Deterministic RGB test master; embedded zones per Spec v1.0",
    }
    return out_path, zonemap


def build_art_04(out_dir: str, w: int, h: int) -> Tuple[str, Dict]:
    """
    ART-04: EDGE + MICROSTRUCTURE MAP (Sharpness / Halos / Noise / Banding)
    Grayscale crisp line composition with line pairs, halo zone, dot grid, radial gradient, edge ring.
    """
    ring_px = int(round(min(w, h) * EDGE_RING_FRAC))
    canvas = make_canvas(w, h, base_y_pct=55.0)

    zones = {}

    # Z1 Line Pair Resolution: 1px,2px,3px separations (vertical + horizontal)
    # We'll draw three panels with alternating lines.
    panel_y0 = int(h * 0.12)
    panel_y1 = int(h * 0.32)
    panel_x0 = int(w * 0.08)
    panel_x1 = int(w * 0.52)
    panel_w = panel_x1 - panel_x0
    gap = int(w * 0.02)
    sep_list = [1, 2, 3]
    panel_each = (panel_w - 2 * gap) // 3
    panels = []
    for i, sep in enumerate(sep_list):
        x0 = panel_x0 + i * (panel_each + gap)
        x1 = x0 + panel_each
        # base fill
        draw_rect_lin(canvas, x0, panel_y0, x1, panel_y1, 60.0)
        # vertical line pairs: black lines with spacing sep
        for x in range(x0, x1, sep * 2):
            canvas[panel_y0:panel_y1, x : min(x1, x + 1)] = 0.0
        # horizontal line pairs: overlay in lower half
        midy = (panel_y0 + panel_y1) // 2
        for y in range(midy, panel_y1, sep * 2):
            canvas[y : min(panel_y1, y + 1), x0:x1] = 0.0
        panels.append({"sep_px": sep, "rect": [x0, panel_y0, x1, panel_y1]})
    zones["Z1"] = {"id": "Z1", "name": "Line Pair Resolution", "panels": panels}

    # Z2 Halo Edge Zone: high-contrast black/white edges
    x0 = int(w * 0.60)
    y0 = int(h * 0.12)
    x1 = int(w * 0.92)
    y1 = int(h * 0.32)
    # left half black, right half white
    mid = (x0 + x1) // 2
    draw_rect_lin(canvas, x0, y0, mid, y1, 0.0)
    draw_rect_lin(canvas, mid, y0, x1, y1, 100.0)
    zones["Z2"] = {"id": "Z2", "name": "Halo Edge Zone", "rect": [x0, y0, x1, y1], "note": "hard edge for halo detection"}

    # Z3 Low-Contrast Dot Grid: 1px grid at 2% contrast
    x0 = int(w * 0.08)
    y0 = int(h * 0.40)
    x1 = int(w * 0.40)
    y1 = int(h * 0.62)
    base = 50.0
    contrast = 2.0
    draw_rect_lin(canvas, x0, y0, x1, y1, base)
    # dots every 6 px
    step = 6
    for yy in range(y0, y1, step):
        for xx in range(x0, x1, step):
            canvas[yy : min(y1, yy + 1), xx : min(x1, xx + 1)] = (base + contrast) / 100.0
    zones["Z3"] = {"id": "Z3", "name": "Low-Contrast Dot Grid", "rect": [x0, y0, x1, y1], "base_y_pct": base, "contrast_y_pct": contrast}

    # Z4 Radial Smooth Gradient (for stepping)
    x0 = int(w * 0.48)
    y0 = int(h * 0.40)
    x1 = int(w * 0.92)
    y1 = int(h * 0.88)
    hh = y1 - y0
    ww = x1 - x0
    yy, xx = np.mgrid[0:hh, 0:ww]
    cx, cy = ww * 0.5, hh * 0.5
    rr = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2) / np.sqrt(cx**2 + cy**2)
    # center 70% down to 35% at edge
    grad = (0.70 - (0.70 - 0.35) * rr).astype(np.float32)
    canvas[y0:y1, x0:x1] = clamp01(grad)
    zones["Z4"] = {"id": "Z4", "name": "Radial Smooth Gradient", "rect": [x0, y0, x1, y1], "y_pct_range": [35, 70]}

    # Z5 Edge ring
    apply_edge_ring_lin(canvas, ring_px, EDGE_RING_Y_PCT)
    zones["Z5"] = {"id": "Z5", "name": "Peripheral Mount Band", "ring_px": ring_px, "y_pct": EDGE_RING_Y_PCT}

    out_path = os.path.join(out_dir, f"ART-04_EdgeMicrostructureMap_v{ARTWORK_VERSION}.png")
    save_png_grayscale(out_path, canvas)

    zonemap = {
        "Artwork_ID": "ART-04",
        "Artwork_Version": ARTWORK_VERSION,
        "ZoneMap_Version": ZONEMAP_VERSION,
        "MasterResolution": {"w": w, "h": h},
        "EdgeRing": {"width_px": ring_px, "y_pct": EDGE_RING_Y_PCT, "srgb8": y_pct_to_srgb8(EDGE_RING_Y_PCT)},
        "Zones": zones,
        "Notes": "Deterministic test master; grayscale; embedded zones per Spec v1.0",
    }
    return out_path, zonemap


ACCESS_SCHEMA_SQL = r"""
' Infusist Experimental Art Suite (D) — Access Schema — Spec v1.0
' NOTE: In Access, run CREATE TABLE statements first, then ALTER TABLE constraints if needed.

CREATE TABLE OptimizerVersion (
  OptimizerVersion_ID TEXT(50) NOT NULL,
  CreatedDate DATETIME,
  Notes MEMO,
  CONSTRAINT PK_OptimizerVersion PRIMARY KEY (OptimizerVersion_ID)
);

CREATE TABLE Artwork (
  Artwork_ID TEXT(20) NOT NULL,
  Artwork_Version TEXT(20),
  ZoneMap_Version TEXT(20),
  Title TEXT(255),
  AspectRatio TEXT(20),
  MasterResolution_LongEdge_px LONG,
  Notes MEMO,
  CONSTRAINT PK_Artwork PRIMARY KEY (Artwork_ID)
);

CREATE TABLE SubstrateBatch (
  SubstrateBatch_ID TEXT(50) NOT NULL,
  Vendor TEXT(255),
  Finish TEXT(50),
  ReceivedDate DATETIME,
  Notes MEMO,
  CONSTRAINT PK_SubstrateBatch PRIMARY KEY (SubstrateBatch_ID)
);

CREATE TABLE Mount (
  Mount_ID TEXT(50) NOT NULL,
  Notes MEMO,
  CONSTRAINT PK_Mount PRIMARY KEY (Mount_ID)
);

CREATE TABLE Operator (
  Operator_ID TEXT(50) NOT NULL,
  Name TEXT(255),
  CONSTRAINT PK_Operator PRIMARY KEY (Operator_ID)
);

CREATE TABLE PrintRun (
  PrintRun_ID TEXT(50) NOT NULL,
  [Date] DATETIME,
  Operator_ID TEXT(50),
  SubstrateBatch_ID TEXT(50),
  Notes MEMO,
  CONSTRAINT PK_PrintRun PRIMARY KEY (PrintRun_ID)
);

CREATE TABLE Print (
  Print_ID TEXT(50) NOT NULL,
  PrintRun_ID TEXT(50),
  Artwork_ID TEXT(20),
  OptimizerVersion_ID TEXT(50),
  Mount_ID TEXT(50),
  Size_Code TEXT(10),
  LongEdge_in DOUBLE,
  ShortEdge_in DOUBLE,
  Viewed_LightCondition TEXT(50),
  PassFail TEXT(10),
  FailureTags TEXT(255),
  Notes MEMO,
  CONSTRAINT PK_Print PRIMARY KEY (Print_ID)
);

CREATE TABLE Observation (
  Observation_ID AUTOINCREMENT PRIMARY KEY,
  Print_ID TEXT(50),
  Lowest_Visible_ShadowStep_pct LONG,
  Shadow_Texture_Visible YESNO,
  Texture_Lowest_Visible_pct LONG,
  Midtone_Banding_Count LONG,
  Highest_Visible_HighlightStep_pct LONG,
  Midtone_Steps_Visible_Count LONG,
  First_Merged_Step_pct LONG,
  MicroContrast_Visible YESNO,
  Ramp_Banding_Count LONG,
  Neutral_Shift TEXT(50),
  Channel_Bias TEXT(10),
  Skin_Shift YESNO,
  Skin_Shift_Type TEXT(50),
  Sat_Steps_Visible_Count LONG,
  Sat_Collapse_Threshold_pct LONG,
  Smallest_Resolved_LinePair TEXT(10),
  Halo_Present YESNO,
  Halo_Width_mm DOUBLE,
  DotGrid_Visible YESNO,
  Gradient_Stepping_Count LONG,
  Edge_Density_Shift TEXT(20),
  Edge_Shift_mm DOUBLE
);

' Foreign keys (Access may require creating relationships in UI depending on settings)
ALTER TABLE PrintRun
  ADD CONSTRAINT FK_PrintRun_Operator FOREIGN KEY (Operator_ID) REFERENCES Operator(Operator_ID);

ALTER TABLE PrintRun
  ADD CONSTRAINT FK_PrintRun_SubstrateBatch FOREIGN KEY (SubstrateBatch_ID) REFERENCES SubstrateBatch(SubstrateBatch_ID);

ALTER TABLE Print
  ADD CONSTRAINT FK_Print_PrintRun FOREIGN KEY (PrintRun_ID) REFERENCES PrintRun(PrintRun_ID);

ALTER TABLE Print
  ADD CONSTRAINT FK_Print_Artwork FOREIGN KEY (Artwork_ID) REFERENCES Artwork(Artwork_ID);

ALTER TABLE Print
  ADD CONSTRAINT FK_Print_OptimizerVersion FOREIGN KEY (OptimizerVersion_ID) REFERENCES OptimizerVersion(OptimizerVersion_ID);

ALTER TABLE Print
  ADD CONSTRAINT FK_Print_Mount FOREIGN KEY (Mount_ID) REFERENCES Mount(Mount_ID);

ALTER TABLE Observation
  ADD CONSTRAINT FK_Observation_Print FOREIGN KEY (Print_ID) REFERENCES Print(Print_ID);

' Recommended indexes
CREATE INDEX idx_Print_Artwork_Opt_Size_Mount
ON Print (Artwork_ID, OptimizerVersion_ID, Size_Code, Mount_ID);

CREATE INDEX idx_Observation_PrintID
ON Observation (Print_ID);
""".strip()


def build_workbook(path: str) -> None:
    wb = Workbook()

    # Sheet 1: README
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "INFUSIST LAB — EXPERIMENTAL ART SUITE (D)"
    ws["A2"] = f"Spec Version: {SUITE_VERSION}"
    ws["A3"] = "Rules: Originals immutable; tests deterministic; no silent drift; no subjective edits."
    ws["A5"] = "Stop conditions:"
    ws["A6"] = "- If any condition yields Failure_Tag rate >10% after first 10 prints -> pause and review"
    ws["A7"] = "- If Failure_Tag rate <5% for all conditions after n=25 -> graduate optimizer version"
    ws["A9"] = "How to use:"
    ws["A10"] = "- Enter print rows in PRINTS"
    ws["A11"] = "- Enter objective measurements in OBSERVATIONS (one row per Print_ID)"
    ws["A12"] = "- QC_SUMMARY is pivot-ready; EXPORT is flattened for Access import"
    ws["A14"] = f"Generator: {GENERATOR_VERSION}"
    ws["A15"] = f"Generated (UTC): {utc_now_iso()}"

    # Sheet 2: LOOKUPS
    lookups = wb.create_sheet("LOOKUPS")
    lists = {
        "FailureTagList": ["TOO_DARK","MIDTONE_CRUSH","HIGHLIGHT_BLOWOUT","LOW_POP","COLOR_CAST","SKIN_SHIFT","NOISE_AMPLIFIED","OVER_SHARPENED","BANDNG_POSTERIZE","SUCCESS"],
        "Neutral_Shift": ["None","Warm","Cool","Green","Magenta","Red","Blue"],
        "Channel_Bias": ["None","R","G","B"],
        "Edge_Density_Shift": ["None","Darker","Lighter"],
        "Skin_Shift_Type": ["TooWarm","TooCool","Green","Magenta"],
        "Smallest_Resolved_LinePair": ["1px","2px","3px","None"],
        "Mount_ID": ["PureShadow","Control"],
        "Size_Code": ["S1","S2"],
        "LightCondition": ["D50","Daylight","WarmLED","CoolLED"],
        "YesNo": ["Y","N"],
        "PassFail": ["Pass","Fail"],
        "Artwork_ID": ["ART-01","ART-02","ART-03","ART-04"],
    }
    row = 1
    for name, vals in lists.items():
        lookups[f"A{row}"] = name
        for i, v in enumerate(vals):
            lookups[f"B{row+i}"] = v
        row += max(2, len(vals) + 1)

    # Helper to define list range
    def list_range(name: str) -> str:
        # find name row
        for r in range(1, lookups.max_row + 1):
            if lookups[f"A{r}"].value == name:
                # values start at B{r}
                rr = r
                # count until blank
                end = r
                while lookups[f"B{end}"].value is not None:
                    end += 1
                end -= 1
                return f"LOOKUPS!$B${rr}:$B${end}"
        raise ValueError(f"Lookup list not found: {name}")

    # Sheet 3: PRINTS
    prints = wb.create_sheet("PRINTS")
    prints_headers = [
        "Print_ID","Date","PrintRun_ID","Artwork_ID","OptimizerVersion_ID","Mount_ID","Size_Code",
        "LongEdge_in","ShortEdge_in","Viewed_LightCondition","PassFail","FailureTags","Notes"
    ]
    prints.append(prints_headers)

    # Add data validations
    dv_artwork = DataValidation(type="list", formula1=f"={list_range('Artwork_ID')}", allow_blank=False)
    dv_mount = DataValidation(type="list", formula1=f"={list_range('Mount_ID')}", allow_blank=False)
    dv_size = DataValidation(type="list", formula1=f"={list_range('Size_Code')}", allow_blank=False)
    dv_light = DataValidation(type="list", formula1=f"={list_range('LightCondition')}", allow_blank=False)
    dv_passfail = DataValidation(type="list", formula1=f"={list_range('PassFail')}", allow_blank=False)

    # Apply validations to a reasonable row range (2..2000)
    max_rows = 2000
    prints.add_data_validation(dv_artwork)
    prints.add_data_validation(dv_mount)
    prints.add_data_validation(dv_size)
    prints.add_data_validation(dv_light)
    prints.add_data_validation(dv_passfail)

    dv_artwork.add(f"D2:D{max_rows}")
    dv_mount.add(f"F2:F{max_rows}")
    dv_size.add(f"G2:G{max_rows}")
    dv_light.add(f"J2:J{max_rows}")
    dv_passfail.add(f"K2:K{max_rows}")

    # FailureTags: custom validation
    # Rule: must include SUCCESS OR one/more tags; SUCCESS cannot co-exist with other tags; not blank.
    # We implement a simple custom formula:
    # - cell not blank
    # - if contains "SUCCESS" then it must equal exactly "SUCCESS"
    #
    # Note: Excel formula uses comma separators; some locales need semicolon. This is a known limitation.
    dv_failtags = DataValidation(type="custom", formula1='=AND(LEN($L2)>0,IF(ISNUMBER(SEARCH("SUCCESS",$L2)),$L2="SUCCESS",TRUE))', allow_blank=False)
    dv_failtags.errorTitle = "Invalid FailureTags"
    dv_failtags.error = 'FailureTags must be "SUCCESS" OR one/more tags separated by "|" (SUCCESS cannot co-exist).'
    prints.add_data_validation(dv_failtags)
    dv_failtags.add(f"L2:L{max_rows}")

    # Sheet 4: OBSERVATIONS
    obs = wb.create_sheet("OBSERVATIONS")
    obs_headers = [
        "Observation_ID","Print_ID","Lowest_Visible_ShadowStep_pct","Shadow_Texture_Visible","Texture_Lowest_Visible_pct",
        "Midtone_Banding_Count","Highest_Visible_HighlightStep_pct","Midtone_Steps_Visible_Count","First_Merged_Step_pct",
        "MicroContrast_Visible","Ramp_Banding_Count","Neutral_Shift","Channel_Bias","Skin_Shift","Skin_Shift_Type",
        "Sat_Steps_Visible_Count","Sat_Collapse_Threshold_pct","Smallest_Resolved_LinePair","Halo_Present","Halo_Width_mm",
        "DotGrid_Visible","Gradient_Stepping_Count","Edge_Density_Shift","Edge_Shift_mm"
    ]
    obs.append(obs_headers)

    dv_yesno = DataValidation(type="list", formula1=f"={list_range('YesNo')}", allow_blank=False)
    dv_neutral = DataValidation(type="list", formula1=f"={list_range('Neutral_Shift')}", allow_blank=False)
    dv_channel = DataValidation(type="list", formula1=f"={list_range('Channel_Bias')}", allow_blank=False)
    dv_skin_type = DataValidation(type="list", formula1=f"={list_range('Skin_Shift_Type')}", allow_blank=True)
    dv_linepair = DataValidation(type="list", formula1=f"={list_range('Smallest_Resolved_LinePair')}", allow_blank=False)
    dv_edge_shift = DataValidation(type="list", formula1=f"={list_range('Edge_Density_Shift')}", allow_blank=False)

    obs.add_data_validation(dv_yesno)
    obs.add_data_validation(dv_neutral)
    obs.add_data_validation(dv_channel)
    obs.add_data_validation(dv_skin_type)
    obs.add_data_validation(dv_linepair)
    obs.add_data_validation(dv_edge_shift)

    # Columns by letter (1-indexed): Shadow_Texture_Visible=D, MicroContrast_Visible=J, Skin_Shift=N, Halo_Present=S, DotGrid_Visible=U
    dv_yesno.add(f"D2:D{max_rows}")  # Shadow_Texture_Visible
    dv_yesno.add(f"J2:J{max_rows}")  # MicroContrast_Visible
    dv_yesno.add(f"N2:N{max_rows}")  # Skin_Shift
    dv_yesno.add(f"S2:S{max_rows}")  # Halo_Present
    dv_yesno.add(f"U2:U{max_rows}")  # DotGrid_Visible

    dv_neutral.add(f"L2:L{max_rows}")
    dv_channel.add(f"M2:M{max_rows}")
    dv_skin_type.add(f"O2:O{max_rows}")
    dv_linepair.add(f"R2:R{max_rows}")
    dv_edge_shift.add(f"W2:W{max_rows}")

    # Numeric bounds:
    # Lowest_Visible_ShadowStep_pct allowed: 0,2,4,6,8,10
    dv_shadow_step = DataValidation(type="list", formula1='"0,2,4,6,8,10"', allow_blank=False)
    obs.add_data_validation(dv_shadow_step)
    dv_shadow_step.add(f"C2:C{max_rows}")

    # Highest_Visible_HighlightStep_pct allowed: 92..100 integers
    dv_highlight = DataValidation(type="whole", operator="between", formula1="92", formula2="100", allow_blank=False)
    obs.add_data_validation(dv_highlight)
    dv_highlight.add(f"G2:G{max_rows}")

    # counts >=0 integer for Midtone_Banding_Count, Ramp_Banding_Count, Gradient_Stepping_Count
    dv_count = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
    obs.add_data_validation(dv_count)
    dv_count.add(f"F2:F{max_rows}")  # Midtone_Banding_Count
    dv_count.add(f"K2:K{max_rows}")  # Ramp_Banding_Count
    dv_count.add(f"V2:V{max_rows}")  # Gradient_Stepping_Count

    # Conditional required fields via custom validation:
    # Texture_Lowest_Visible_pct required if Shadow_Texture_Visible="Y"
    dv_tex_req = DataValidation(type="custom", formula1='=IF($D2="Y",LEN($E2)>0,TRUE)', allow_blank=True)
    dv_tex_req.errorTitle = "Required field"
    dv_tex_req.error = "Texture_Lowest_Visible_pct required when Shadow_Texture_Visible=Y"
    obs.add_data_validation(dv_tex_req)
    dv_tex_req.add(f"E2:E{max_rows}")

    # Halo_Width_mm required if Halo_Present="Y"
    dv_halo_req = DataValidation(type="custom", formula1='=IF($S2="Y",LEN($T2)>0,TRUE)', allow_blank=True)
    dv_halo_req.errorTitle = "Required field"
    dv_halo_req.error = "Halo_Width_mm required when Halo_Present=Y"
    obs.add_data_validation(dv_halo_req)
    dv_halo_req.add(f"T2:T{max_rows}")

    # Sheet 5: QC_SUMMARY (pivot-ready headings)
    qc = wb.create_sheet("QC_SUMMARY")
    qc_headers = [
        "Artwork_ID","OptimizerVersion_ID","Size_Code","Mount_ID",
        "Total Prints","Fail Count","Fail Rate",
        "TOO_DARK","MIDTONE_CRUSH","HIGHLIGHT_BLOWOUT","LOW_POP","COLOR_CAST","SKIN_SHIFT","NOISE_AMPLIFIED","OVER_SHARPENED","BANDNG_POSTERIZE",
        "Median Lowest_Visible_ShadowStep (Pivot)","Median Highest_Visible_HighlightStep (Pivot)","Halo Rate (Pivot)"
    ]
    qc.append(qc_headers)
    qc["A2"] = "Use Excel PivotTables on EXPORT sheet for summary + medians."

    # Sheet 6: CONTROL_LIMITS
    cl = wb.create_sheet("CONTROL_LIMITS")
    cl["A1"] = "Control limits (Spec v1.0)"
    cl["A3"] = "Shadow floor target <=4"
    cl["A4"] = "Highlight ceiling target >=97"
    cl["A5"] = "FailRate target <5%"
    cl["A6"] = "Halo width target <1mm"

    # Sheet 7: EXPORT (flattened)
    ex = wb.create_sheet("EXPORT")
    export_headers = prints_headers + obs_headers[1:]  # exclude Observation_ID; include Print_ID onward
    ex.append(export_headers)

    # Map EXPORT row N to PRINTS row N and OBSERVATIONS row N by Print_ID lookup.
    # Deterministic formula approach:
    # - Copy PRINTS fields directly by row (A..M)
    # - For OBSERVATIONS: XLOOKUP Print_ID into OBSERVATIONS!B:B and return each field
    #
    # If older Excel: replace XLOOKUP with INDEX/MATCH. We use XLOOKUP for simplicity.
    # Users can pivot from EXPORT.
    for r in range(2, max_rows + 1):
        # PRINTS copy
        for c in range(1, len(prints_headers) + 1):
            col = get_column_letter(c)
            ex[f"{col}{r}"] = f"=PRINTS!{col}{r}"

        # OBSERVATIONS pull by Print_ID (EXPORT col N onward)
        # Determine start column for obs fields:
        start_c = len(prints_headers) + 1
        # OBSERVATIONS columns (from Print_ID onward):
        obs_fields = obs_headers[1:]  # includes Print_ID first
        # We skip Print_ID because already in PRINTS col A.
        # For each obs field after Print_ID, xlookup.
        for i, field in enumerate(obs_fields[1:], start=0):
            dest_col = get_column_letter(start_c + i)
            # Find the column index in OBSERVATIONS sheet for this field
            src_idx = obs_headers.index(field) + 1  # 1-based
            src_col = get_column_letter(src_idx)
            # XLOOKUP: lookup Print_ID in OBSERVATIONS!B:B, return OBSERVATIONS!<src_col>:<src_col>
            ex[f"{dest_col}{r}"] = f'=IFERROR(XLOOKUP($A{r},OBSERVATIONS!$B:$B,OBSERVATIONS!${src_col}:${src_col},""),"")'

    # Column widths (basic)
    for ws2 in [prints, obs, ex]:
        ws2.freeze_panes = "A2"
        for i in range(1, min(30, ws2.max_column) + 1):
            ws2.column_dimensions[get_column_letter(i)].width = 18

    wb.save(path)


def write_json(path: str, obj: Dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, sort_keys=True)


def write_text(path: str, text: str) -> None:
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(text)


def zip_dir(zip_path: str, base_dir: str) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(base_dir):
            for fn in files:
                fp = os.path.join(root, fn)
                rel = os.path.relpath(fp, base_dir)
                z.write(fp, rel)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out_dir", default="out_suite", help="Output directory (will be created)")
    ap.add_argument("--w", type=int, default=DEFAULT_W, help="Master width px")
    ap.add_argument("--h", type=int, default=DEFAULT_H, help="Master height px")
    ap.add_argument("--zip", action="store_true", help="Also create ZIP bundle")
    args = ap.parse_args()

    out_dir = os.path.abspath(args.out_dir)
    ensure_dir(out_dir)

    settings = {
        "generator_version": GENERATOR_VERSION,
        "suite_version": SUITE_VERSION,
        "artwork_version": ARTWORK_VERSION,
        "zonemap_version": ZONEMAP_VERSION,
        "master_resolution": {"w": args.w, "h": args.h},
        "edge_ring_frac": EDGE_RING_FRAC,
        "edge_ring_y_pct": EDGE_RING_Y_PCT,
        "mapping": {
            "y_pct_to_srgb8": "linear_light_pct -> sRGB gamma -> 8-bit",
            "note": "Deterministic; for test art reproducibility, not full ICC workflow."
        },
        "generated_utc": utc_now_iso(),
    }

    # Build artworks + zonemaps
    manifest_files: List[Dict] = []
    def add_file(p: str, kind: str) -> None:
        manifest_files.append({"path": os.path.relpath(p, out_dir), "kind": kind, "sha256": sha256_file(p)})

    art1_path, z1 = build_art_01(out_dir, args.w, args.h)
    z1_path = os.path.join(out_dir, f"ZoneMap_ART-01_v{ZONEMAP_VERSION}.json")
    write_json(z1_path, z1)
    add_file(art1_path, "artwork_png")
    add_file(z1_path, "zonemap_json")

    art2_path, z2 = build_art_02(out_dir, args.w, args.h)
    z2_path = os.path.join(out_dir, f"ZoneMap_ART-02_v{ZONEMAP_VERSION}.json")
    write_json(z2_path, z2)
    add_file(art2_path, "artwork_png")
    add_file(z2_path, "zonemap_json")

    art3_path, z3 = build_art_03(out_dir, args.w, args.h)
    z3_path = os.path.join(out_dir, f"ZoneMap_ART-03_v{ZONEMAP_VERSION}.json")
    write_json(z3_path, z3)
    add_file(art3_path, "artwork_png")
    add_file(z3_path, "zonemap_json")

    art4_path, z4 = build_art_04(out_dir, args.w, args.h)
    z4_path = os.path.join(out_dir, f"ZoneMap_ART-04_v{ZONEMAP_VERSION}.json")
    write_json(z4_path, z4)
    add_file(art4_path, "artwork_png")
    add_file(z4_path, "zonemap_json")

    # Workbook
    wb_path = os.path.join(out_dir, f"InfusistLab_Workbook_ExpArtSuiteD_v{SUITE_VERSION}.xlsx")
    build_workbook(wb_path)
    add_file(wb_path, "xlsx_workbook")

    # Access schema
    sql_path = os.path.join(out_dir, f"InfusistLab_AccessSchema_ExpArtSuiteD_v{SUITE_VERSION}.sql")
    write_text(sql_path, ACCESS_SCHEMA_SQL)
    add_file(sql_path, "access_schema_sql")

    # Manifest
    manifest = {
        "suite": "Infusist Experimental Art Suite (D)",
        "spec_version": SUITE_VERSION,
        "settings_json": settings,
        "files": sorted(manifest_files, key=lambda x: x["path"]),
    }
    manifest_path = os.path.join(out_dir, "manifest.json")
    write_json(manifest_path, manifest)
    add_file(manifest_path, "manifest")

    # ZIP bundle
    if args.zip:
        zip_path = os.path.join(os.path.dirname(out_dir), f"Infusist_ExpArtSuiteD_v{SUITE_VERSION}_EXPORT.zip")
        zip_dir(zip_path, out_dir)
        # hash zip too
        with open(zip_path + ".sha256", "w", encoding="utf-8", newline="\n") as f:
            f.write(f"{sha256_file(zip_path)}  {os.path.basename(zip_path)}\n")

    print("BUILD OK")
    print(f"OUT_DIR: {out_dir}")
    print("FILES:")
    for f in manifest["files"]:
        print(f" - {f['path']}  {f['sha256']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())